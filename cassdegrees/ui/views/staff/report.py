from django.shortcuts import render
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse

from api.models import CourseModel, ProgramModel, SubplanModel
from api.views import search

from openpyxl import Workbook
from openpyxl.styles import Font

import json


def generate_course_info_table():
    # Retrieve all courses and store them in a list.
    all_courses = CourseModel.objects.all()
    all_courses_list = list()

    # Key: course code,
    # Value: dictionary with key as program and/or subplan, value as list of programs/subplans dependent on course.
    all_dependencies = dict()

    # Convert all QuerySet results to Python dictionaries.
    for course in all_courses:
        all_courses_list.append(model_to_dict(course))

    # Sort all courses by order of course codes.
    all_courses_list = sorted(all_courses_list, key=lambda i: i['code'])

    # Generate an internal request to search api made by Jack
    gen_request = HttpRequest()

    # Determine all the subplans and programs that depends on any courses.
    for course in all_courses_list:
        gen_request.GET = {'select': 'code,year,name,rules', 'from': 'subplan', 'rules': course['code']}
        # subplans which depend on course where its code is equal to course['code']
        subplans = json.loads(search(gen_request).content.decode())
        gen_request.GET = {'select': 'code,year,name,rules', 'from': 'program', 'rules': course['code']}
        # programs which depend on course where its code is equal to course['code']
        programs = json.loads(search(gen_request).content.decode())
        all_dependencies[course['code']] = {
            'subplans': subplans,
            'programs': programs
        }

    table_headings = [
        "Code",
        "Name",
        "Units",
        "Years Offered",
        "Offerings",
        "Subplan Dependencies",
        "Program Dependencies",
        "Status"
    ]

    table_body = list()

    for course in all_courses_list:
        new_row = list()
        new_row.append(course.get("code"))
        new_row.append(course.get("name"))
        new_row.append(course.get("units"))
        new_row.append(course.get("offeredYears"))

        # Build a string for the "Offering" column
        offerings = list()
        if course.get("offeredSem1"):
            offerings.append("Sem1 ")
        if course.get("offeredSem2"):
            offerings.append("Sem2 ")
        if course.get("offeredSummer"):
            offerings.append("Summer ")
        if course.get("offeredAutumn"):
            offerings.append("Autumn ")
        if course.get("offeredWinter"):
            offerings.append("Winter ")
        if course.get("offeredSpring"):
            offerings.append("Spring ")
        if course.get("otherOffering"):
            offerings.append("Other ")

        offerings = ", ".join(offerings)
        new_row.append(offerings)

        # Build a string for the "Subplan Dependency" column
        subplan_dependencies = list()

        for subplan in all_dependencies.get(course.get("code")).get("subplans"):
            subplan_dependencies.append(subplan.get("code"))

        subplan_dependencies = ", ".join(subplan_dependencies)
        new_row.append(subplan_dependencies)

        # Build a string for the "Program Dependency" column
        program_dependencies = list()

        for program in all_dependencies.get(course.get("code")).get("programs"):
            program_dependencies.append(program.get("code"))

        program_dependencies = ", ".join(program_dependencies)
        new_row.append(program_dependencies)

        # Add the right value to the "Status" column
        if course.get("currentlyActive"):
            new_row.append("Active")
        else:
            new_row.append("Inactive")

        table_body.append(new_row)

    return table_headings, table_body


# Generates the courses report and then uses that to generate an excel file based off that.
# https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
def generate_excel(request):
    response = HttpResponse(content_type='application/ms-excel')

    # File name is defined here
    response['Content-Disposition'] = 'attachment; filename="Course Report.xlsx"'

    wb = Workbook()  # Create new workbook
    ws = wb.active
    ws.title = 'Courses'

    columns, body = generate_course_info_table()

    row_ind = 1  # Openpyxl worksheets start at index 1 for rows and columns.

    for col_ind in range(len(columns)):
        cell = ws.cell(row=row_ind, column=col_ind + 1)
        cell.value = columns[col_ind]
        cell.font = Font(bold=True)  # Make header rows bold

    # Write all the body rows to the excel workbook.
    for row in body:
        row_ind += 1

        for col_ind in range(len(row)):
            cell = ws.cell(row=row_ind, column=col_ind + 1)
            cell.value = row[col_ind]

    wb.save(response)
    return response


# Simply generates and displays report of all courses in web view.
def report_section(request):
    table_headings, table_body = generate_course_info_table()

    context = {
        'table_headings': table_headings,
        'table_body': table_body
    }

    return render(request, 'staff/report/coursereport.html', context=context)
